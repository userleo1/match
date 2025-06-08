import sys
import uuid
from PyQt5.QtWidgets import (QMainWindow, QAction, QFileDialog, QMessageBox, QSplitter,
                             QListView, QTableWidget, QTableWidgetItem, QPushButton,
                             QVBoxLayout, QWidget, QLineEdit, QComboBox, QDialog,
                             QHBoxLayout, QLabel, QProgressBar, QMenu, QInputDialog)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QStandardItemModel, QStandardItem
import pandas as pd
import openpyxl
import pymysql
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import hashlib
import re


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.current_project = None
        self.current_table_model = None
        self.projects = {}  # 存储已打开的工程
        self.quota_tables = []  # 存储数据库中的定额表

    def initUI(self):
        super().__init__()
        # 设置窗口标题和大小
        self.setWindowTitle('定额匹配软件')
        self.setGeometry(100, 100, 1200, 800)

        # 创建菜单栏
        self.createMenus()

        # 创建中心部件
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)

        # 创建分割器，左侧为工程列表，右侧为匹配表格
        self.splitter = QSplitter(Qt.Horizontal)

        # 左侧工程列表
        self.project_list_view = QListView()
        self.project_model = QStandardItemModel()
        self.project_list_view.setModel(self.project_model)
        self.project_list_view.clicked.connect(self.onProjectSelected)
        self.splitter.addWidget(self.project_list_view)
        self.project_list_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.project_list_view.customContextMenuRequested.connect(self.showProjectContextMenu)
        self.project_list_view.setDragEnabled(True)
        self.project_list_view.setAcceptDrops(True)
        self.project_list_view.setDropIndicatorShown(True)
        self.project_list_view.setDragDropMode(QListView.InternalMove)
        # 右侧匹配表格区域
        self.right_widget = QWidget()
        self.right_layout = QVBoxLayout(self.right_widget)

        # 匹配表格
        self.match_table = QTableWidget()
        self.match_table.setEditTriggers(QTableWidget.AllEditTriggers)
        self.right_layout.addWidget(self.match_table)

        # 按钮区域
        self.button_layout = QHBoxLayout()
        self.start_match_button = QPushButton('开始匹配')
        self.start_match_button.clicked.connect(self.onStartMatch)
        self.fix_button = QPushButton('修正')
        self.fix_button.clicked.connect(self.onFix)
        self.button_layout.addWidget(self.start_match_button)
        self.button_layout.addWidget(self.fix_button)
        self.right_layout.addLayout(self.button_layout)

        self.splitter.addWidget(self.right_widget)
        self.splitter.setSizes([200, 1000])  # 初始大小

        self.main_layout.addWidget(self.splitter)

        # 连接数据库，获取定额表列表
        self.refreshQuotaTables()

    def createMenus(self):
        # 创建菜单栏
        menubar = self.menuBar()

        # 文件菜单
        file_menu = menubar.addMenu('文件')

        # 定额导入模板
        template_action = QAction('定额导入模板', self)
        template_action.setShortcut('Ctrl+T')
        template_action.triggered.connect(self.generateQuotaTemplate)
        file_menu.addAction(template_action)

        # 导入定额
        import_quota_action = QAction('导入定额', self)
        import_quota_action.setShortcut('Ctrl+I')
        import_quota_action.triggered.connect(self.importQuota)
        file_menu.addAction(import_quota_action)

        # 新建工程
        new_project_action = QAction('新建工程', self)
        new_project_action.setShortcut('Ctrl+N')
        new_project_action.triggered.connect(self.createNewProject)
        file_menu.addAction(new_project_action)

        # 打开工程
        open_project_action = QAction('打开工程', self)
        open_project_action.setShortcut('Ctrl+O')
        open_project_action.triggered.connect(self.openProject)
        file_menu.addAction(open_project_action)

        # 保存
        save_action = QAction('保存', self)
        save_action.setShortcut('Ctrl+S')
        save_action.triggered.connect(self.saveProject)
        file_menu.addAction(save_action)

        # 另存为
        save_as_action = QAction('另存为', self)
        save_as_action.setShortcut('Ctrl+Shift+S')
        save_as_action.triggered.connect(self.saveProjectAs)
        file_menu.addAction(save_as_action)

        # 关闭
        close_action = QAction('关闭', self)
        close_action.setShortcut('Ctrl+W')
        close_action.triggered.connect(self.closeProject)
        file_menu.addAction(close_action)

        # 匹配菜单
        match_menu = menubar.addMenu('匹配')

        # 导出匹配文件模板
        export_match_template_action = QAction('导出匹配文件模板', self)
        export_match_template_action.triggered.connect(self.exportMatchTemplate)
        match_menu.addAction(export_match_template_action)

        # 导入匹配文件
        import_match_action = QAction('导入匹配文件', self)
        import_match_action.triggered.connect(self.importMatchFile)
        match_menu.addAction(import_match_action)

        # 开始匹配文件
        start_match_menu_action = QAction('开始匹配文件', self)
        start_match_menu_action.triggered.connect(self.onStartMatch)
        match_menu.addAction(start_match_menu_action)

        # 导出匹配结果
        export_match_result_action = QAction('导出匹配结果', self)
        export_match_result_action.triggered.connect(self.exportMatchResult)
        match_menu.addAction(export_match_result_action)

        # 帮助菜单
        help_menu = menubar.addMenu('帮助')

        # 说明
        instruction_action = QAction('说明', self)
        instruction_action.triggered.connect(self.showInstructions)
        help_menu.addAction(instruction_action)

        # 关于我们
        about_action = QAction('关于我们', self)
        about_action.triggered.connect(self.showAbout)
        help_menu.addAction(about_action)

    def refreshQuotaTables(self):
        """刷新定额表列表"""
        try:
            # 先连接到MySQL服务器（不指定数据库）
            conn = pymysql.connect(
                host='localhost',
                user='root',
                password='L4512678.1'
            )
            cursor = conn.cursor()

            # 检查数据库是否存在，不存在则创建
            cursor.execute("SHOW DATABASES LIKE 'coast'")
            if not cursor.fetchone():
                cursor.execute("CREATE DATABASE coast CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")

            # 现在连接到coast数据库
            conn = pymysql.connect(
                host='localhost',
                user='root',
                password='L4512678.1',
                database='coast'
            )
            cursor = conn.cursor()

            # 获取所有表
            cursor.execute("SHOW TABLES")
            tables = cursor.fetchall()

            # 过滤出不是绑定表的表（绑定表格式为xxx_bind）
            self.quota_tables = [table[0] for table in tables if not table[0].endswith('_bind')]
            conn.close()
        except Exception as e:
            QMessageBox.critical(self, '数据库错误', f'连接数据库失败: {str(e)}')

    def generateQuotaTemplate(self):
        """生成定额导入模板"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, '保存模板', '', 'Excel 文件 (*.xlsx)'
            )
            if not file_path:
                return

            # 创建Excel文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '定额模板'

            # 设置表头
            headers = [
                '定额编号', '分部分项工程名称', '计量单位', '工程量',
                '主材费', '小计', '人工费', '材料费', '机械费'
            ]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

            # 保存文件
            wb.save(file_path)
            QMessageBox.information(self, '成功', '定额导入模板已生成')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'生成模板失败: {str(e)}')

    def importQuota(self):
        """导入定额数据"""
        print("进入importQuota函数")  # 调试输出
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, '选择定额文件', '', 'Excel 文件 (*.xlsx *.xls)'
            )

            if not file_path:
                return

            # 读取Excel文件
            df = pd.read_excel(file_path)

            df.info()  # 打印数据基本信息

            print(df.head().to_dict(orient='records'))  # 打印数据前几行

            # 检查文件是否为空
            if df.empty:
                QMessageBox.critical(self, '错误', '导入的文件没有数据')
                return

            # 改进的模板验证逻辑
            required_headers = [
                '定额编号', '分部分项工程名称', '计量单位', '工程量',
                '主材费', '小计', '人工费', '材料费', '机械费'
            ]

            # 检查列数是否一致
            if len(df.columns) != len(required_headers):
                error_msg = f'列数不匹配，模板需要{len(required_headers)}列，实际{len(df.columns)}列'
                QMessageBox.critical(self, '错误', error_msg)
                print(error_msg)  # 调试输出
                print(f"模板列名: {required_headers}")
                print(f"文件列名: {list(df.columns)}")
                return

            # 检查每一列是否存在且名称正确
            missing_headers = []
            for header in required_headers:
                if header not in df.columns:
                    missing_headers.append(header)

            if missing_headers:
                error_msg = f'缺少必要的列: {", ".join(missing_headers)}'
                QMessageBox.critical(self, '错误', error_msg)
                return

            # 更严格的列名匹配（位置和内容都要匹配）
            for i, (expected, actual) in enumerate(zip(required_headers, df.columns)):
                if expected.strip() != actual.strip():
                    error_msg = f'第{i + 1}列列名不匹配: 期望 "{expected}"，实际 "{actual}"'
                    QMessageBox.critical(self, '错误', error_msg)
                    return

            # 处理数据中的NaN值
            df = df.fillna('')  # 将NaN值替换为空字符串

            # 获取表名（文件名去掉扩展名）
            original_table_name = file_path.split('/')[-1].split('.')[0]
            # 替换表名中的非法字符，但保留中文
            safe_table_name = re.sub(r'[^\w\u4e00-\u9fa5]', '_', original_table_name)

            # 连接数据库
            conn = pymysql.connect(
                host='localhost',
                user='root',
                password='L4512678.1',
                database='coast',
                charset='utf8mb4',  # 指定字符集支持中文
                autocommit=False  # 关闭自动提交
            )
            cursor = conn.cursor()

            # 检查表是否存在
            cursor.execute(f"SHOW TABLES LIKE '{safe_table_name}'")
            table_exists = cursor.fetchone()

            if table_exists:
                # 检查表中是否已有数据
                cursor.execute(f"SELECT COUNT(*) FROM {safe_table_name}")
                count = cursor.fetchone()[0]

                if count > 0:
                    # 检查导入的数据是否与已有数据重复
                    first_code = df.iloc[0]['定额编号']
                    cursor.execute(f"SELECT COUNT(*) FROM {safe_table_name} WHERE `定额编号` = %s", (first_code,))
                    code_count = cursor.fetchone()[0]

                    if code_count > 0:
                        QMessageBox.critical(self, '错误', '定额已存在')
                        conn.close()
                        return

            # 创建表（如果不存在）
            if not table_exists:
                # 生成创建表的SQL语句
                create_table_sql = f"CREATE TABLE IF NOT EXISTS `{safe_table_name}` ("  # 使用反引号包裹表名
                create_table_sql += "id INT AUTO_INCREMENT PRIMARY KEY, "

                # 添加列，使用反引号包裹列名
                for col in df.columns:
                    # 替换列名中的非法字符，但保留中文
                    safe_col = re.sub(r'[^\w\u4e00-\u9fa5]', '_', col)
                    create_table_sql += f"`{safe_col}` TEXT, "

                # 移除最后一个逗号和空格
                create_table_sql = create_table_sql[:-2]
                create_table_sql += ")"

                cursor.execute(create_table_sql)

            # 保存数据库连接和表名的引用，防止被垃圾回收
            self.import_conn = conn
            self.import_table_name = safe_table_name

            # 插入数据
            # 使用多线程处理，避免界面卡顿
            class ImportThread(QThread):
                progress_updated = pyqtSignal(int)
                import_completed = pyqtSignal(bool, str)

                def __init__(self, conn, table_name, df):
                    super().__init__()
                    self.conn = conn
                    self.table_name = table_name
                    self.df = df

                def run(self):
                    print(f"导入线程开始处理表 {self.table_name}")  # 调试输出
                    try:
                        cursor = self.conn.cursor()
                        total_rows = len(self.df)

                        # 构建插入语句
                        # 使用反引号包裹列名
                        columns = ', '.join([f"`{re.sub(r'[^\w\u4e00-\u9fa5]', '_', col)}`" for col in self.df.columns])
                        placeholders = ', '.join(['%s'] * len(self.df.columns))
                        insert_sql = f"INSERT INTO `{self.table_name}` ({columns}) VALUES ({placeholders})"

                        # 批量插入
                        for i, row in enumerate(self.df.itertuples(index=False)):
                            values = tuple(row)

                            # 打印前几行数据用于调试
                            if i < 5:
                                print(f"插入第{i + 1}行数据: {values}")  # 调试输出

                            try:
                                cursor.execute(insert_sql, values)
                            except Exception as e:
                                # 打印单行插入失败的具体错误
                                print(f"第{i + 1}行插入失败: {str(e)}")
                                print(f"失败数据: {values}")
                                raise  # 继续抛出异常，终止导入过程

                            # 更新进度
                            if i % 10 == 0:  # 每10行更新一次进度
                                progress = int((i + 1) / total_rows * 100)
                                self.progress_updated.emit(progress)
                                print(f"导入进度: {progress}%")  # 调试输出

                        self.conn.commit()
                        self.import_completed.emit(True, '导入成功')
                        print(f"表 {self.table_name} 导入成功")  # 调试输出
                    except Exception as e:
                        self.conn.rollback()
                        error_msg = f'导入失败: {str(e)}'
                        self.import_completed.emit(False, error_msg)
                        import traceback
                        print(f"导入失败: {error_msg}")  # 调试输出
                        print(traceback.format_exc())  # 打印详细的错误堆栈信息
                    finally:
                        # 确保数据库连接在导入完成后不会被关闭
                        print(f"导入线程处理表 {self.table_name} 结束")  # 调试输出

            # 创建进度对话框
            progress_dialog = QDialog(self)
            progress_dialog.setWindowTitle('导入进度')
            progress_dialog.setMinimumWidth(400)
            progress_layout = QVBoxLayout(progress_dialog)

            progress_label = QLabel('正在导入数据...')
            progress_layout.addWidget(progress_label)

            progress_bar = QProgressBar()
            progress_bar.setRange(0, 100)
            progress_layout.addWidget(progress_bar)

            # 创建导入线程
            import_thread = ImportThread(conn, safe_table_name, df)
            import_thread.progress_updated.connect(progress_bar.setValue)
            import_thread.import_completed.connect(
                lambda success, msg: self.onImportCompleted(success, msg, progress_dialog, conn))

            # 显示进度对话框并启动线程
            progress_dialog.show()
            print("启动导入线程...")  # 调试输出
            import_thread.start()

            # 保持对线程的引用，防止被垃圾回收
            self.current_import_thread = import_thread

        except Exception as e:
            QMessageBox.critical(self, '错误', f'导入定额失败: {str(e)}')
            import traceback
            print(f"全局异常捕获: {str(e)}")  # 调试输出
            print(traceback.format_exc())  # 打印详细的错误堆栈信息

    def onImportCompleted(self, success, message, dialog, conn):
        """导入完成后的回调函数"""
        dialog.close()
        conn.close()

        if success:
            QMessageBox.information(self, '成功', message)
            # 刷新定额表列表
            self.refreshQuotaTables()
        else:
            QMessageBox.critical(self, '错误', message)

    def createNewProject(self):
        """创建新工程"""
        try:
            # 检查数据库连接
            conn = pymysql.connect(
                host='localhost',
                user='root',
                password='L4512678.1',
                database='coast'
            )
            cursor = conn.cursor()

            # 检查定额表是否存在
            cursor.execute("SHOW TABLES")
            tables = [table[0] for table in cursor.fetchall()]
            self.quota_tables = [table for table in tables]

            if not self.quota_tables:
                QMessageBox.critical(self, '错误', '请先导入定额表')
                conn.close()
                return

            # 创建对话框
            dialog = QDialog(self)
            dialog.setWindowTitle('新建工程')
            dialog.setMinimumWidth(400)

            layout = QVBoxLayout(dialog)

            # 工程名称
            name_layout = QHBoxLayout()
            name_label = QLabel('工程名称:')
            name_edit = QLineEdit()
            name_layout.addWidget(name_label)
            name_layout.addWidget(name_edit)
            layout.addLayout(name_layout)

            # 选择定额表
            table_layout = QHBoxLayout()
            table_label = QLabel('选择定额表:')
            table_combo = QComboBox()
            table_combo.addItems(self.quota_tables)
            table_layout.addWidget(table_label)
            table_layout.addWidget(table_combo)
            layout.addLayout(table_layout)

            # 按钮
            button_layout = QHBoxLayout()
            ok_button = QPushButton('确定')
            cancel_button = QPushButton('取消')
            button_layout.addWidget(ok_button)
            button_layout.addWidget(cancel_button)
            layout.addLayout(button_layout)

            # 连接信号
            ok_button.clicked.connect(dialog.accept)
            cancel_button.clicked.connect(dialog.reject)

            if dialog.exec_():
                project_name = name_edit.text().strip()

                # 检查工程名是否为空
                if not project_name:
                    QMessageBox.critical(self, '错误', '工程名称不能为空')
                    conn.close()
                    return

                # 检查同名工程是否存在
                for project_id, project in self.projects.items():
                    if project['name'] == project_name:
                        QMessageBox.critical(self, '错误', '已存在同名工程')
                        conn.close()
                        return

                selected_table = table_combo.currentText()

                # 创建绑定表
                bind_table_name = f"{selected_table}_bind"

                # 总是创建绑定表(IF NOT EXISTS)
                cursor.execute(f"SHOW COLUMNS FROM {selected_table}")
                columns = cursor.fetchall()

                # 创建绑定表SQL
                create_bind_table_sql = f"CREATE TABLE IF NOT EXISTS `{bind_table_name}` ("
                create_bind_table_sql += "`id` INT AUTO_INCREMENT PRIMARY KEY, "
                create_bind_table_sql += "`project_id` VARCHAR(255) NOT NULL, "
                create_bind_table_sql += "`condition_hash` CHAR(32) NOT NULL, "

                # 添加定额表的列
                for col in columns:
                    col_name = col[0]
                    if col_name != 'id':  # 跳过id列
                        create_bind_table_sql += f"`{col_name}` TEXT, "

                # 添加use_count列
                create_bind_table_sql += "`use_count` INT DEFAULT 0, "

                # 添加唯一索引
                create_bind_table_sql += "UNIQUE KEY `project_condition` (`project_id`, `condition_hash`)"
                create_bind_table_sql += ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"

                cursor.execute(create_bind_table_sql)
                print(f"创建绑定表SQL: {create_bind_table_sql}")  # 调试输出

                conn.close()

                # 创建工程对象
                project = {
                    'name': project_name,
                    'id': str(uuid.uuid4()),
                    'quota_table': selected_table,
                    'bind_table': bind_table_name,
                    'match_file': None,
                    'match_data': None,
                    'table_has_data': True  # 不再检查表格数据
                }

                # 添加到工程列表
                self.projects[project['id']] = project

                # 添加到UI列表
                item = QStandardItem(project_name)
                item.setData(project['id'])
                self.project_model.appendRow(item)

                # 设置为当前工程
                self.current_project = project['id']
                self.onProjectSelected()

                QMessageBox.information(self, '成功', '工程创建成功')

        except pymysql.Error as e:
            QMessageBox.critical(self, '数据库错误', f'连接数据库失败: {str(e)}')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'创建工程失败: {str(e)}')

    def openProject(self):
        """打开工程"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, '打开工程', '', '工程文件 (*.yq)'
            )
            if not file_path:
                return

            # 读取工程文件
            with open(file_path, 'r') as f:
                import json
                project_data = json.load(f)

                # 检查工程数据完整性
                required_fields = ['name', 'id', 'quota_table', 'bind_table']
                for field in required_fields:
                    if field not in project_data:
                        QMessageBox.critical(self, '错误', '工程文件格式不正确')
                        return

                # 添加到工程列表
                self.projects[project_data['id']] = project_data

                # 添加到UI列表
                item = QStandardItem(project_data['name'])
                item.setData(project_data['id'])
                self.project_model.appendRow(item)

                # 设置为当前工程
                self.current_project = project_data['id']
                self.onProjectSelected()

                QMessageBox.information(self, '成功', '工程已打开')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'打开工程失败: {str(e)}')

    def saveProject(self):
        """保存工程"""
        if not self.current_project:
            QMessageBox.critical(self, '错误', '没有打开的工程')
            return

        project = self.projects[self.current_project]

        # 如果没有工程文件路径，调用另存为
        if 'file_path' not in project:
            self.saveProjectAs()
            return

        try:
            # 保存工程数据
            with open(project['file_path'], 'w') as f:
                import json
                json.dump(project, f)

            QMessageBox.information(self, '成功', '工程已保存')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'保存工程失败: {str(e)}')

    def saveProjectAs(self):
        """另存为工程"""
        if not self.current_project:
            QMessageBox.critical(self, '错误', '没有打开的工程')
            return

        project = self.projects[self.current_project]

        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, '保存工程', '', '工程文件 (*.yq)'
            )
            if not file_path:
                return

            # 添加文件路径到工程数据
            project['file_path'] = file_path

            # 保存工程数据
            with open(file_path, 'w') as f:
                import json
                json.dump(project, f)

            QMessageBox.information(self, '成功', '工程已另存为')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'另存为工程失败: {str(e)}')

    def closeProject(self):
        """关闭当前工程"""
        if not self.current_project:
            QMessageBox.critical(self, '错误', '没有打开的工程')
            return

        # 从UI列表中移除
        for i in range(self.project_model.rowCount()):
            item = self.project_model.item(i)
            if item.data() == self.current_project:
                self.project_model.removeRow(i)
                break

        # 从工程列表中移除
        del self.projects[self.current_project]

        # 清空当前工程
        self.current_project = None

        # 清空表格
        self.match_table.setRowCount(0)
        self.match_table.setColumnCount(0)

        QMessageBox.information(self, '成功', '工程已关闭')

    def exportMatchTemplate(self):
        """导出匹配文件模板"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, '保存模板', '', 'Excel 文件 (*.xlsx)'
            )
            if not file_path:
                return

            # 创建Excel文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '匹配模板'

            # 设置表头
            headers = ['序号', '名称', '规格', '型号', '工作内容', '项目特征', '修正项目']
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

            # 保存文件
            wb.save(file_path)
            QMessageBox.information(self, '成功', '匹配文件模板已生成')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'生成模板失败: {str(e)}')

    def importMatchFile(self):
        """导入匹配文件"""
        if not self.current_project:
            QMessageBox.critical(self, '错误', '请先打开或创建工程')
            return

        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, '选择匹配文件', '', 'Excel 文件 (*.xlsx *.xls)'
            )
            if not file_path:
                return

            # 读取Excel文件
            df = pd.read_excel(file_path)

            # 验证模板
            required_headers = ['序号', '名称', '规格', '型号', '工作内容', '项目特征', '修正项目']

            if list(df.columns) != required_headers:
                QMessageBox.critical(self, '错误', '请使用模板导入')
                return

            # 更新工程信息
            project = self.projects[self.current_project]
            project['match_file'] = file_path
            project['match_data'] = df.to_dict('records')

            # 显示数据到表格
            self.displayMatchData(df)

            QMessageBox.information(self, '成功', '匹配文件已导入')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导入匹配文件失败: {str(e)}')

    def displayMatchData(self, df):
        # 过滤序号列（保留之前修改）
        display_df = df.drop(columns=['序号']) if '序号' in df.columns else df

        # 清空表格
        self.match_table.setRowCount(0)
        self.match_table.setColumnCount(0)

        # 设置列数和列标题
        self.match_table.setColumnCount(len(display_df.columns))
        self.match_table.setHorizontalHeaderLabels(display_df.columns)

        # 设置数据
        for row_idx, row_data in enumerate(display_df.itertuples(index=False)):
            self.match_table.insertRow(row_idx)
            for col_idx, value in enumerate(row_data):
                # 保留之前的文本处理逻辑
                cell_text = str(value) if pd.notna(value) else ''
                # 增强正则：匹配>=3空格或>=1制表符（保留之前换行处理）
                processed_text = re.sub(r'(\s{3,}|\t+)', '\n', cell_text)

                item = QTableWidgetItem(processed_text)
                item.setTextAlignment(Qt.AlignTop | Qt.TextWordWrap)  # 保持自动换行

                # 列名判断可编辑性（保留序号过滤后的列索引处理）
                column_name = display_df.columns[col_idx]
                if column_name in ['规格', '型号', '修正项目']:
                    item.setFlags(item.flags() | Qt.ItemIsEditable)
                else:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                self.match_table.setItem(row_idx, col_idx, item)

        # 保持UI调整（保留所有优化）
        self.match_table.resizeColumnsToContents()
        self.match_table.resizeRowsToContents()

        # 设置最小行高
        for row in range(self.match_table.rowCount()):
            self.match_table.verticalHeader().setDefaultSectionSize(40)
        # 在resizeRowsToContents()后添加
        for row in range(self.match_table.rowCount()):
            height = 0
            for col in range(self.match_table.columnCount()):
                item = self.match_table.item(row, col)
                if item:
                    height = max(height, self.match_table.sizeHintForIndex(
                        self.match_table.model().index(row, col)).height())
            if height > 0:
                self.match_table.setRowHeight(row, height + 10)  # 增加10像素边距

    def showProjectContextMenu(self, pos):
        """显示工程上下文菜单"""
        index = self.project_list_view.indexAt(pos)
        if index.isValid():
            menu = QMenu()
            rename_action = QAction('重命名', self)
            delete_action = QAction('删除', self)
            menu.addAction(rename_action)
            menu.addAction(delete_action)

            rename_action.triggered.connect(lambda: self.renameProject(index))
            delete_action.triggered.connect(lambda: self.deleteProject(index))
            menu.exec_(self.project_list_view.viewport().mapToGlobal(pos))

    def renameProject(self, index):
        """重命名工程"""
        item = self.project_model.itemFromIndex(index)
        old_name = item.text()
        new_name, ok = QInputDialog.getText(
            self, '重命名', '输入新名称:', text=old_name
        )
        if ok and new_name:
            item.setText(new_name)
            project_id = item.data()
            self.projects[project_id]['name'] = new_name

    def deleteProject(self, index):
        """删除工程"""
        reply = QMessageBox.question(
            self, '确认', '确定要删除该工程吗？',
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            item = self.project_model.itemFromIndex(index)
            project_id = item.data()

            # 从数据结构中删除
            del self.projects[project_id]
            # 从模型中删除
            self.project_model.removeRow(index.row())

        # 修改后的 onProjectSelected 方法

    def onProjectSelected(self):
        """工程选择事件处理"""
        indexes = self.project_list_view.selectedIndexes()
        if not indexes:
            return

        index = indexes[0]
        item = self.project_model.itemFromIndex(index)
        self.current_project = item.data()

        # 更新右侧显示
        project = self.projects.get(self.current_project)
        if project and 'match_data' in project and project['match_data']:
            df = pd.DataFrame(project['match_data'])
            self.displayMatchData(df)
        else:
            self.match_table.setRowCount(0)
            self.match_table.setColumnCount(0)

    def onStartMatch(self):
        """开始匹配"""
        if not self.current_project:
            QMessageBox.critical(self, '错误', '请先打开或创建工程')
            return

        project = self.projects[self.current_project]

        # 检查是否已导入匹配文件
        if 'match_data' not in project or not project['match_data']:
            QMessageBox.critical(self, '错误', '请先导入匹配文件')
            return

        # 获取匹配数据
        match_data = project['match_data']

        if not match_data:
            QMessageBox.critical(self, '错误', '没有可匹配的数据')
            return

        # 创建匹配线程
        class MatchThread(QThread):
            progress_updated = pyqtSignal(int)
            match_completed = pyqtSignal(bool, str)
            match_result = pyqtSignal(list)

            def __init__(self, project, match_data):
                super().__init__()
                self.project = project
                self.match_data = match_data

            def run(self):
                try:
                    conn = pymysql.connect(
                        host='localhost',
                        user='root',
                        password='L4512678.1',
                        database='coast'
                    )
                    cursor = conn.cursor()

                    total_rows = len(self.match_data)
                    result = []

                    # 获取定额表结构
                    cursor.execute(f"SHOW COLUMNS FROM {self.project['quota_table']}")
                    quota_columns = [col[0] for col in cursor.fetchall()]

                    # 对于每一行匹配数据
                    for i, row in enumerate(self.match_data):
                        # 更新进度
                        if i % 10 == 0:  # 每10行更新一次进度
                            progress = int((i + 1) / total_rows * 100)
                            self.progress_updated.emit(progress)

                        # 复制当前行数据
                        result_row = row.copy()

                        # 如果修正项目已有编码，直接使用
                        if row.get('修正项目'):
                            code = row['修正项目']

                            # 从定额表中查询该编码的详细信息
                            query = f"SELECT * FROM {self.project['quota_table']} WHERE 定额编号 = %s"
                            cursor.execute(query, (code,))
                            quota_data = cursor.fetchone()

                            if quota_data:
                                # 将定额数据添加到结果中
                                for col_idx, col_name in enumerate(quota_columns):
                                    if col_name != 'id':  # 跳过id列
                                        result_row[f'定额_{col_name}'] = quota_data[col_idx]
                        else:
                            # 计算匹配条件的哈希值
                            condition = f"{row.get('名称', '')}{row.get('规格', '')}{row.get('型号', '')}{row.get('工作内容', '')}{row.get('项目特征', '')}"
                            condition_hash = hashlib.md5(condition.encode('utf-8')).hexdigest()

                            # 先从绑定表中查找
                            bind_query = f"SELECT * FROM {self.project['bind_table']} WHERE project_id = %s AND condition_hash = %s ORDER BY use_count DESC LIMIT 1"
                            cursor.execute(bind_query, (self.project['id'], condition_hash))
                            bind_data = cursor.fetchone()

                            if bind_data:
                                # 使用绑定表中的编码
                                code = bind_data[quota_columns.index('定额编号')]  # 假设定额编号在第2列
                                result_row['修正项目'] = code

                                # 更新使用次数
                                update_query = f"UPDATE {self.project['bind_table']} SET use_count = use_count + 1 WHERE id = %s"
                                cursor.execute(update_query, (bind_data[0],))
                                conn.commit()

                                # 将定额数据添加到结果中
                                for col_idx, col_name in enumerate(quota_columns):
                                    if col_name != 'id':  # 跳过id列
                                        result_row[f'定额_{col_name}'] = bind_data[
                                            col_idx + 3]  # 绑定表中多了project_id, condition_hash, use_count三列
                            else:
                                # 没有绑定数据，使用文本匹配
                                # 拼接匹配条件
                                match_text = f"{row.get('名称', '')} {row.get('规格', '')} {row.get('型号', '')} {row.get('工作内容', '')} {row.get('项目特征', '')}"

                                # 获取所有定额数据
                                cursor.execute(f"SELECT * FROM {self.project['quota_table']}")
                                quota_rows = cursor.fetchall()

                                # 计算相似度
                                best_similarity = -1
                                best_code = None
                                best_quota_data = None

                                for quota_row in quota_rows:
                                    # 拼接定额文本
                                    quota_text = f"{quota_row[1]} {quota_row[2]} {quota_row[3]} {quota_row[4]} {quota_row[5]}"  # 假设前5列是名称、规格、型号、工作内容、项目特征

                                    # 计算相似度
                                    similarity = self.calculate_similarity(match_text, quota_text)

                                    if similarity > best_similarity:
                                        best_similarity = similarity
                                        best_code = quota_row[1]  # 假设定额编号在第2列
                                        best_quota_data = quota_row

                                if best_code:
                                    result_row['修正项目'] = best_code

                                    # 将定额数据添加到结果中
                                    for col_idx, col_name in enumerate(quota_columns):
                                        if col_name != 'id':  # 跳过id列
                                            result_row[f'定额_{col_name}'] = best_quota_data[col_idx]

                        result.append(result_row)

                    conn.close()
                    self.match_result.emit(result)
                    self.match_completed.emit(True, '匹配完成')
                except Exception as e:
                    self.match_completed.emit(False, str(e))

            def calculate_similarity(self, text1, text2):
                """计算两个文本的相似度"""
                if not text1 or not text2:
                    return 0

                # 使用简单的余弦相似度
                vectorizer = TfidfVectorizer()
                tfidf_matrix = vectorizer.fit_transform([text1, text2])
                similarity = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]

                return similarity

        # 创建进度对话框
        progress_dialog = QDialog(self)
        progress_dialog.setWindowTitle('匹配进度')
        progress_dialog.setMinimumWidth(400)
        progress_layout = QVBoxLayout(progress_dialog)

        progress_label = QLabel('正在匹配...')
        progress_layout.addWidget(progress_label)

        progress_bar = QProgressBar()
        progress_bar.setRange(0, 100)
        progress_layout.addWidget(progress_bar)

        # 创建匹配线程
        match_thread = MatchThread(project, match_data)
        match_thread.progress_updated.connect(progress_bar.setValue)
        match_thread.match_completed.connect(lambda success, msg: self.onMatchCompleted(success, msg, progress_dialog))
        match_thread.match_result.connect(self.onMatchResult)

        # 显示进度对话框并启动线程
        progress_dialog.show()
        match_thread.start()

    def onMatchCompleted(self, success, message, dialog):
        """匹配完成后的回调函数"""
        dialog.close()

        if success:
            QMessageBox.information(self, '成功', message)
        else:
            QMessageBox.critical(self, '错误', message)

    def onMatchResult(self, result):
        """处理匹配结果"""
        if not self.current_project:
            return

        # 更新工程数据
        project = self.projects[self.current_project]
        project['match_data'] = result

        # 显示结果
        df = pd.DataFrame(result)
        self.displayMatchData(df)

    def onFix(self):
        """修正按钮点击事件"""
        if not self.current_project:
            QMessageBox.critical(self, '错误', '请先打开或创建工程')
            return

        project = self.projects[self.current_project]

        # 获取当前选中的行
        selected_rows = set()
        for item in self.match_table.selectedItems():
            selected_rows.add(item.row())

        if not selected_rows:
            QMessageBox.critical(self, '错误', '请选择要修正的数据行')
            return

        try:
            conn = pymysql.connect(
                host='localhost',
                user='root',
                password='L4512678.1',
                database='coast'
            )
            cursor = conn.cursor()

            # 处理每一行
            for row_idx in selected_rows:
                # 获取匹配条件和修正后的编码
                name = self.match_table.item(row_idx, 1).text() if self.match_table.item(row_idx, 1) else ''
                specification = self.match_table.item(row_idx, 2).text() if self.match_table.item(row_idx, 2) else ''
                model = self.match_table.item(row_idx, 3).text() if self.match_table.item(row_idx, 3) else ''
                work_content = self.match_table.item(row_idx, 4).text() if self.match_table.item(row_idx, 4) else ''
                project_feature = self.match_table.item(row_idx, 5).text() if self.match_table.item(row_idx, 5) else ''
                code = self.match_table.item(row_idx, 6).text() if self.match_table.item(row_idx, 6) else ''

                if not code:
                    continue  # 没有修正编码，跳过

                # 计算条件哈希
                condition = f"{name}{specification}{model}{work_content}{project_feature}"
                condition_hash = hashlib.md5(condition.encode('utf-8')).hexdigest()

                # 检查表中是否已有该条件的记录
                check_query = f"SELECT id FROM {project['bind_table']} WHERE project_id = %s AND condition_hash = %s"
                cursor.execute(check_query, (project['id'], condition_hash))
                existing_record = cursor.fetchone()

                if existing_record:
                    # 更新记录
                    update_query = f"UPDATE {project['bind_table']} SET use_count = use_count + 1 WHERE id = %s"
                    cursor.execute(update_query, (existing_record[0],))
                else:
                    # 插入新记录
                    # 获取定额表结构
                    cursor.execute(f"SHOW COLUMNS FROM {project['quota_table']}")
                    quota_columns = [col[0] for col in cursor.fetchall()]

                    # 从定额表中查询该编码的详细信息
                    query = f"SELECT * FROM {project['quota_table']} WHERE 定额编号 = %s"
                    cursor.execute(query, (code,))
                    quota_data = cursor.fetchone()

                    if quota_data:
                        # 构建插入语句
                        insert_columns = "project_id, condition_hash, "
                        insert_values = "%s, %s, "
                        insert_params = [project['id'], condition_hash]

                        for col_idx, col_name in enumerate(quota_columns):
                            if col_name != 'id':  # 跳过id列
                                insert_columns += f"`{col_name}`, "
                                insert_values += "%s, "
                                insert_params.append(quota_data[col_idx])

                        # 添加use_count
                        insert_columns += "use_count"
                        insert_values += "1"

                        insert_query = f"INSERT INTO {project['bind_table']} ({insert_columns}) VALUES ({insert_values})"
                        cursor.execute(insert_query, insert_params)

            conn.commit()
            conn.close()

            QMessageBox.information(self, '成功', '修正已保存')
        except Exception as e:
            conn.rollback()
            conn.close()
            QMessageBox.critical(self, '错误', f'保存修正失败: {str(e)}')

    def exportMatchResult(self):
        """导出匹配结果"""
        if not self.current_project:
            QMessageBox.critical(self, '错误', '请先打开或创建工程')
            return

        project = self.projects[self.current_project]

        # 检查是否有匹配数据
        if 'match_data' not in project or not project['match_data']:
            QMessageBox.critical(self, '错误', '没有匹配数据可导出')
            return

        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, '保存匹配结果', '', 'Excel 文件 (*.xlsx)'
            )
            if not file_path:
                return

            # 创建Excel文件
            df = pd.DataFrame(project['match_data'])
            df.to_excel(file_path, index=False)

            QMessageBox.information(self, '成功', '匹配结果已导出')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导出匹配结果失败: {str(e)}')

    def showInstructions(self):
        """显示使用说明"""
        QMessageBox.information(
            self,
            '使用说明',
            '定额匹配软件使用说明：\n\n'
            '1. 首先导入定额数据：\n'
            '   - 点击"文件"菜单，选择"定额导入模板"生成导入模板\n'
            '   - 按照模板填写定额数据\n'
            '   - 点击"文件"菜单，选择"导入定额"导入填写好的定额数据\n\n'
            '2. 创建新工程：\n'
            '   - 点击"文件"菜单，选择"新建工程"\n'
            '   - 输入工程名称，选择要使用的定额表\n\n'
            '3. 导入匹配文件：\n'
            '   - 点击"匹配"菜单，选择"导出匹配文件模板"生成匹配模板\n'
            '   - 按照模板填写需要匹配的项目\n'
            '   - 点击"匹配"菜单，选择"导入匹配文件"导入填写好的匹配文件\n\n'
            '4. 开始匹配：\n'
            '   - 点击"匹配"菜单，选择"开始匹配文件"或点击工具栏上的"开始匹配"按钮\n'
            '   - 系统会自动将匹配文件中的项目与定额表进行匹配\n\n'
            '5. 修正匹配结果：\n'
            '   - 对于匹配不准确的项目，可以手动修改"修正项目"列中的定额编号\n'
            '   - 点击"修正"按钮保存修正结果\n\n'
            '6. 导出匹配结果：\n'
            '   - 点击"匹配"菜单，选择"导出匹配结果"将匹配结果导出为Excel文件'
        )

    def showAbout(self):
        """显示关于信息"""
        QMessageBox.information(
            self,
            '关于我们',
            '定额匹配软件 v1.0\n\n'
            '联系方式：\n'
            '电话：123-456-7890\n'
            '邮箱：contact@example.com\n'
            '网址：https://www.example.com'
        )